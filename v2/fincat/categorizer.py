"""
AI-powered transaction categorization using Claude.
"""

import json
import logging
import os
import time
from pathlib import Path
from typing import List, Dict

import anthropic
import openpyxl

logger = logging.getLogger('fincat.categorizer')


class Categorizer:
    """Categorize transactions using Claude AI."""

    def __init__(self, config: dict):
        """
        Initialize categorizer.

        Args:
            config: Configuration dictionary
        """
        self.config = config

        # Initialize Anthropic client
        api_key = os.getenv('ANTHROPIC_API_KEY')
        if not api_key:
            raise ValueError("ANTHROPIC_API_KEY not set")

        self.client = anthropic.Anthropic(api_key=api_key)
        self.model = config['ai']['model']
        self.batch_size = config['ai']['batch_size']
        self.max_retries = config['ai']['max_retries']

        # Load categories
        self.categories = self._load_categories()

    def _load_categories(self) -> List[str]:
        """Load categories from reference file, create if doesn't exist."""
        data_folder = Path(self.config['folders']['data'])
        data_folder.mkdir(parents=True, exist_ok=True)

        categories_file = data_folder / self.config['excel']['categories_file']

        if not categories_file.exists():
            logger.info(f"Creating default categories file: {categories_file}")
            self._create_default_categories(categories_file)

        # Read categories from Excel
        wb = openpyxl.load_workbook(categories_file)
        ws = wb.active

        categories = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Category name in first column
                categories.append(str(row[0]).strip())

        logger.info(f"Loaded {len(categories)} categories")
        return categories

    def _create_default_categories(self, filepath: Path):
        """Create default categories file."""
        wb = openpyxl.Workbook()
        ws = wb.active

        # Headers
        ws.append(['קטגוריה', 'תיאור', 'English'])

        # Default categories
        default_categories = [
            ('מזון וסופרמרקט', 'סופרים, מכולת, שוק', 'Food & Groceries'),
            ('מסעדות ובתי קפה', 'אוכל מחוץ לבית', 'Restaurants & Cafes'),
            ('תחבורה ודלק', 'דלק, חניה, תחבורה ציבורית', 'Transportation & Fuel'),
            ('בריאות ורפואה', 'רופאים, תרופות, ביטוח בריאות', 'Health & Medical'),
            ('קניות וביגוד', 'בגדים, נעליים, קניות אונליין', 'Shopping & Clothing'),
            ('בידור ופנאי', 'קולנוע, ספורט, טיולים', 'Entertainment & Leisure'),
            ('חשבונות ושירותים', 'חשמל, מים, אינטרנט, טלפון', 'Bills & Utilities'),
            ('חינוך', 'שכר לימוד, ספרים, קורסים', 'Education'),
            ('ביטוח', 'ביטוחים שונים', 'Insurance'),
            ('אחר', 'לא מוגדר', 'Other')
        ]

        for category in default_categories:
            ws.append(category)

        wb.save(filepath)

    def categorize_transactions(self, transactions: List) -> Dict[str, str]:
        """
        Categorize all transactions.

        Args:
            transactions: List of Transaction objects

        Returns:
            Dictionary mapping business name to category
        """
        # Get unique business names
        business_names = list(set(t.business_name for t in transactions))

        logger.info(f"Categorizing {len(business_names)} unique businesses")

        all_categories = {}

        # Process in batches
        for i in range(0, len(business_names), self.batch_size):
            batch = business_names[i:i + self.batch_size]
            logger.debug(f"Processing batch {i//self.batch_size + 1} ({len(batch)} businesses)")

            batch_categories = self._categorize_batch(batch)
            all_categories.update(batch_categories)

        return all_categories

    def _categorize_batch(self, business_names: List[str]) -> Dict[str, str]:
        """
        Categorize a batch of business names using Claude API.

        Args:
            business_names: List of business names

        Returns:
            Dictionary mapping business name to category
        """
        prompt = self._build_prompt(business_names)

        response_text = self._call_api_with_retry(prompt)

        if response_text:
            return self._parse_response(response_text, business_names)
        else:
            # API failed, return uncategorized
            logger.warning("API failed, marking all as uncategorized")
            return {name: "לא סווג" for name in business_names}

    def _build_prompt(self, business_names: List[str]) -> str:
        """Build categorization prompt for Claude."""
        categories_text = "\n".join(f"- {cat}" for cat in self.categories)

        businesses_text = "\n".join(
            f"{i+1}. {name}" for i, name in enumerate(business_names)
        )

        return f"""You are categorizing Israeli credit card transactions.

Available categories (Hebrew):
{categories_text}

Transaction business names (Hebrew):
{businesses_text}

Instructions:
1. Match each business to the most appropriate category
2. Consider the business type, not just the name
3. Use "אחר" (Other) only if you're uncertain
4. Return valid JSON only, no additional text

Return JSON mapping each business name to its category.
Format: {{"business_name": "category_name"}}
Example: {{"רמי לוי": "מזון וסופרמרקט", "קפה גרג": "מסעדות ובתי קפה"}}
"""

    def _call_api_with_retry(self, prompt: str) -> str:
        """
        Call Claude API with exponential backoff retry.

        Args:
            prompt: Prompt to send to API

        Returns:
            Response text or None if failed
        """
        for attempt in range(self.max_retries):
            try:
                response = self.client.messages.create(
                    model=self.model,
                    max_tokens=1024,
                    temperature=0,
                    messages=[{"role": "user", "content": prompt}]
                )

                # Log token usage for cost tracking
                logger.info(
                    f"API call successful: {response.usage.input_tokens} in, "
                    f"{response.usage.output_tokens} out tokens"
                )

                return response.content[0].text

            except anthropic.APIError as e:
                if attempt == self.max_retries - 1:
                    logger.error(f"API failed after {self.max_retries} attempts: {e}")
                    return None

                wait_time = 2 ** attempt  # 1s, 2s, 4s
                logger.warning(
                    f"API error, retrying in {wait_time}s... "
                    f"(attempt {attempt+1}/{self.max_retries}): {e}"
                )
                time.sleep(wait_time)

        return None

    def _parse_response(self, response_text: str, business_names: List[str]) -> Dict[str, str]:
        """
        Parse JSON response from Claude.

        Args:
            response_text: Response from API
            business_names: Original business names

        Returns:
            Dictionary mapping business name to category
        """
        try:
            # Extract JSON from response (might have markdown code blocks)
            if '```json' in response_text:
                response_text = response_text.split('```json')[1].split('```')[0]
            elif '```' in response_text:
                response_text = response_text.split('```')[1].split('```')[0]

            categories = json.loads(response_text.strip())

            # Validate categories
            valid_categories = set(self.categories)
            for business, category in categories.items():
                if category not in valid_categories:
                    logger.warning(
                        f"Invalid category '{category}' for '{business}', using 'אחר'"
                    )
                    categories[business] = 'אחר'

            # Fill in any missing businesses
            for business in business_names:
                if business not in categories:
                    logger.warning(f"Missing category for '{business}', using 'לא סווג'")
                    categories[business] = 'לא סווג'

            return categories

        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse API response as JSON: {e}")
            logger.debug(f"Response text: {response_text[:200]}")
            return {name: "לא סווג" for name in business_names}
