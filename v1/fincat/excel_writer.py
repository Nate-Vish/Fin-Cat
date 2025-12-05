"""
Master Excel file management.
"""

import logging
import time
from datetime import datetime
from pathlib import Path
from typing import List, Dict

import openpyxl

logger = logging.getLogger('fincat.excel_writer')


class FileLockedError(Exception):
    """Raised when Excel file is locked (open in Excel)."""
    pass


class ExcelWriter:
    """Manage master Excel tracking file."""

    def __init__(self, config: dict):
        """
        Initialize Excel writer.

        Args:
            config: Configuration dictionary
        """
        self.config = config

        data_folder = Path(config['folders']['data'])
        data_folder.mkdir(parents=True, exist_ok=True)

        self.master_file = data_folder / config['excel']['master_file']
        self.lock_wait = config['excel']['file_lock_wait']

    def append_transactions(self, transactions: List, categories: Dict[str, str]):
        """
        Append transactions to master file.

        Args:
            transactions: List of Transaction objects
            categories: Dictionary mapping business name to category
        """
        # Wait for file to be available
        if not self._wait_for_file_available():
            raise FileLockedError(
                f"Master file '{self.master_file.name}' is locked (Excel is open). "
                f"Please close Excel and try again."
            )

        # Load or create workbook
        if self.master_file.exists():
            logger.debug(f"Loading existing master file: {self.master_file}")
            wb = openpyxl.load_workbook(self.master_file)
            ws = wb.active
        else:
            logger.info(f"Creating new master file: {self.master_file}")
            wb, ws = self._create_new_workbook()

        # Append transactions
        for transaction in transactions:
            category = categories.get(transaction.business_name, 'לא סווג')

            row = [
                transaction.date.strftime('%d/%m/%Y'),
                transaction.card,
                transaction.business_name,
                transaction.amount,
                transaction.currency,
                category,
                transaction.installments,
                transaction.source_filename,
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                transaction.details
            ]

            ws.append(row)

        # Save
        wb.save(self.master_file)
        logger.info(f"Appended {len(transactions)} transactions to {self.master_file}")

    def _create_new_workbook(self):
        """Create new master file with headers."""
        wb = openpyxl.Workbook()
        ws = wb.active

        # Headers
        headers = [
            'תאריך',           # Date
            'כרטיס',           # Card
            'שם העסק',         # Business Name
            'סכום',            # Amount
            'מטבע',            # Currency
            'קטגוריה',         # Category
            'תשלומים',         # Installments
            'מקור',            # Source
            'תאריך עיבוד',     # Processing Date
            'הערות'            # Notes
        ]

        ws.append(headers)

        # Make headers bold
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        return wb, ws

    def _is_file_locked(self) -> bool:
        """
        Check if Excel file is currently open (locked).

        Returns:
            True if locked, False if available
        """
        if not self.master_file.exists():
            return False

        try:
            # Try to open in append mode
            with open(self.master_file, 'a'):
                return False
        except (IOError, PermissionError):
            return True

    def _wait_for_file_available(self) -> bool:
        """
        Wait for file to become available (not locked).

        Returns:
            True if file available, False if timeout
        """
        if not self._is_file_locked():
            return True

        logger.warning(
            f"Master file is locked (Excel is open). "
            f"Waiting up to {self.lock_wait} seconds..."
        )

        for i in range(self.lock_wait):
            if not self._is_file_locked():
                logger.info("File is now available")
                return True

            if i % 5 == 0 and i > 0:
                logger.info(f"Still waiting... ({i}/{self.lock_wait}s)")

            time.sleep(1)

        return False
